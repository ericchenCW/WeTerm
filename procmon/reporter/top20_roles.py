"""TOP 20 资源占用榜单 + 蓝鲸角色对号。

复用 procmon_report 的 loader/aggregate/normalize，按服务（cmdline_key）聚合后，
对每个 (host, key) 推断它属于哪个蓝鲸角色，最后按内存 / CPU / IO 各出一张
全主机汇总的 TOP 20。

用法：
    cd reporter && python3 top20_roles.py \
        --data-dir "../data/办公网ITOM性能数据_2026.5.25下午~5.26上午" \
        --host-meta ../data/host-meta.json
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import pandas as pd

from procmon_report.aggregate import aggregate_by_service, compute_deltas, summarize
from procmon_report.loader import load_dir
from procmon_report.normalize import normalize_key

# 蓝鲸已知角色全集（来自 host-meta 的 roles 字段汇总）。括号子角色拍平成主角色。
KNOWN_ROLES = [
    "iam", "ssm", "usermgr", "gse", "license", "redis", "consul", "nfs",
    "es7", "monitorv3", "nginx", "rabbitmq", "appo", "appt", "influxdb",
    "mongodb", "zk", "kafka", "paas", "cmdb", "job", "nodeman", "log",
]

# 蓝鲸部署目录名 → host-meta 角色名 的别名表。蓝鲸的 .envs / 工作目录普遍带
# bk 前缀（bkmonitorv3 / bknodeman），和 host-meta 里的简称（monitorv3 / nodeman）
# 对不上，这里统一拉齐。paas_agent 是 appo（应用引擎-正式）的工作目录。
ROLE_ALIAS = {
    "bkmonitorv3": "monitorv3", "bkmonitor": "monitorv3", "monitorv3": "monitorv3",
    "bknodeman": "nodeman", "nodeman": "nodeman",
    "bkpaas": "paas", "paas": "paas", "open_paas": "paas",
    "paas_agent": "appo",
    "usermgr": "usermgr",
    "bkiam": "iam", "iam": "iam",
    "bklog": "log", "log": "log",
    "bkssm": "ssm", "ssm": "ssm",
    "bkcmdb": "cmdb", "cmdb": "cmdb",
    "bkjob": "job", "job": "job",
    "bkgse": "gse", "gse": "gse",
    "bklicense": "license", "license": "license",
}

# comm（进程名）到蓝鲸角色的直接映射 —— 用于中间件类进程，它们的角色名
# 不出现在部署路径里，只能靠进程名/cmdline 特征识别。
COMM_ROLE = {
    "redis-server": "redis",
    "redis-sentinel": "redis",
    "consul": "consul",
    "mongod": "mongodb",
    "mongos": "mongodb",
    "nginx": "nginx",
    "influxd": "influxdb",
    "beam.smp": "rabbitmq",      # erlang VM = rabbitmq
    "rabbitmq-server": "rabbitmq",
    "grafana": "monitorv3",       # 默认归 monitorv3，log(grafana) 后面用路径修正
    "grafana-server": "monitorv3",
}

# 非蓝鲸角色的系统/杂项进程，单独归一类，避免污染榜单的角色解读。
SYSTEM_COMMS = {"systemd", "crond", "logrotate", "rsyslogd", "sshd", "agetty",
                "polkitd", "dbus-daemon", "auditd", "chronyd", "tuned",
                "systemd-journal", "systemd-logind"}

# 第三方/本地运维 agent，非蓝鲸角色，单独归“其他”。
OTHER_COMMS = {"titanagent", "perfect_oom_mon", "vector", "node_exporter",
               "filebeat", "telegraf"}


def _seg_to_role(seg: str) -> str | None:
    seg = seg.split("-")[0]               # .envs/usermgr-api → usermgr
    if seg in ROLE_ALIAS:
        return ROLE_ALIAS[seg]
    if seg in KNOWN_ROLES:
        return seg
    return None


def infer_role(comm: str, cmdline: str, cwd: str, container: str | None,
               host_roles: set[str]) -> tuple[str, str]:
    """返回 (role, 依据)。host_roles 是该主机声明承载的角色集合，用来约束/消歧。"""
    comm = (comm or "").strip()
    cmdline = (cmdline or "").strip()
    cwd = (cwd or "").strip()
    container = (container or "").strip()
    blob = f"{cwd} {cmdline}"

    # 0) PaaS 应用引擎托管的 SaaS 容器：cwd=/data/app/code 或 /cache/.bk，由 appo 拉起。
    #    role 归 appo（host-meta 里的角色名），容器名在服务列已能体现是哪个 SaaS。
    if container:
        if container == "vector":
            return "其他", "vector 日志采集"
        if "/data/app/code" in cwd or "/cache/.bk" in blob or cwd == "/app":
            return "appo", f"SaaS 容器 {container}"
        return "appo", f"容器 {container}"

    # 1) 部署路径 /data/bkce/<role>/... 或 .envs/<role>-... 最可靠（含别名拉齐）
    for m in re.finditer(r"/data/bkce/(?:\.envs/)?([a-z0-9_]+)", blob):
        role = _seg_to_role(m.group(1))
        if role:
            return role, f"路径 /data/bkce/{m.group(1)}"
        if m.group(1) in ("open", "apigw", "bkapi"):
            return "paas", f"路径段 {m.group(1)}"

    # 2) DB 备份任务（cron 触发的临时进程，非常驻角色）
    if "/backup/" in blob or "dbbak" in blob or comm in ("mongodump", "mysqldump"):
        return "备份", "dbbak 备份任务"

    # 3) zookeeper / kafka 的 java 进程靠 cmdline 主类识别
    if "zookeeper" in blob.lower() or "QuorumPeerMain" in cmdline:
        return "zk", "cmdline=zookeeper"
    if "kafka.Kafka" in cmdline or "/kafka/" in blob.lower():
        return "kafka", "cmdline=kafka"

    # 4) elasticsearch java 进程 → es7
    if "elasticsearch" in blob.lower():
        return "es7", "cmdline=elasticsearch"

    # 5) gse agent / proc
    if "/gse/" in cwd.lower() or "/gse" in cmdline.lower() or comm.startswith("gse"):
        return "gse", "gse 路径/进程名"

    # 6) comm 直接映射的中间件
    if comm in COMM_ROLE:
        role = COMM_ROLE[comm]
        if role == "monitorv3" and "log" in cwd.lower():
            return "log", "grafana@log"
        return role, f"进程名 {comm}"

    # 7) 系统 / 第三方运维 agent
    if comm in SYSTEM_COMMS:
        return "系统", f"系统进程 {comm}"
    if comm in OTHER_COMMS or comm.replace(".", "") in OTHER_COMMS:
        return "其他", f"运维 agent {comm}"

    return "未匹配", ""


def export_excel(summary: pd.DataFrame, path: Path, top_n: int) -> None:
    """导出 xlsx：三张 TOP-N 榜单 + 一张全量明细，均带角色列。"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    GB = 1024 * 1024
    MB = 1024 * 1024

    def board(metric: str):
        t = summary.sort_values(metric, ascending=False).head(top_n).reset_index(drop=True)
        return pd.DataFrame({
            "排名": range(1, len(t) + 1),
            "蓝鲸角色": t["role"],
            "主机": t["host"],
            "服务 (cmdline_key)": t["key"],
            "内存峰值(GB)": (t["rss_kb_peak"] / GB).round(2),
            "CPU峰值(%)": t["cpu_pct_peak"].round(0),
            "IO总量(GB)": (t["io_total_bytes"] / 1024 / GB).round(2),
            "对号依据": t["why"],
        })

    boards = {
        "TOP20-内存": board("rss_kb_peak"),
        "TOP20-CPU": board("cpu_pct_peak"),
        "TOP20-IO": board("io_total_bytes"),
    }

    detail = summary.sort_values("rss_kb_peak", ascending=False).copy()
    detail_df = pd.DataFrame({
        "蓝鲸角色": detail["role"],
        "对号依据": detail["why"],
        "主机": detail["host"],
        "服务 (cmdline_key)": detail["key"],
        "采样数": detail["samples"],
        "内存均值(GB)": (detail["rss_kb_avg"] / GB).round(2),
        "内存P95(GB)": (detail["rss_kb_p95"] / GB).round(2),
        "内存峰值(GB)": (detail["rss_kb_peak"] / GB).round(2),
        "内存趋势(%)": detail["rss_trend_pct"].round(1),
        "CPU均值(%)": detail["cpu_pct_avg"].round(0),
        "CPU P95(%)": detail["cpu_pct_p95"].round(0),
        "CPU峰值(%)": detail["cpu_pct_peak"].round(0),
        "IO总量(GB)": (detail["io_total_bytes"] / 1024 / GB).round(2),
        "IO均速(MB/s)": (detail["io_rate_avg"] / MB).round(2),
        "IO峰速(MB/s)": (detail["io_rate_peak"] / MB).round(2),
    })

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, frame in {**boards, "全量明细": detail_df}.items():
            frame.to_excel(xw, sheet_name=name, index=False)
            ws = xw.sheets[name]
            ws.freeze_panes = "A2"
            head_fill = PatternFill("solid", fgColor="305496")
            for col_idx, col in enumerate(frame.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = head_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                width = max(len(str(col)) * 1.6,
                            *(len(str(v)) for v in frame[col].astype(str)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 8), 42)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", required=True, type=Path)
    ap.add_argument("--host-meta", required=True, type=Path)
    ap.add_argument("--top-n", type=int, default=20)
    ap.add_argument("--excel", type=Path, default=None,
                    help="同时导出 xlsx（三张榜单 + 全量明细）")
    args = ap.parse_args()

    meta = json.loads(args.host_meta.read_text(encoding="utf-8"))
    host_roles_map = {
        h: {r.split("(")[0].strip() for r in spec["roles"].split(",")}
        for h, spec in meta["hosts"].items()
    }

    df, stats = load_dir(args.data_dir)
    print(f"loaded files={stats.files} lines={stats.lines} bad={stats.bad_lines}")

    df["cmdline_key"] = df.apply(
        lambda r: normalize_key(str(r["comm"]), str(r["cmdline"]),
                                cwd=r.get("cwd"), container=r.get("container")),
        axis=1,
    )
    df = compute_deltas(df)
    df_agg = aggregate_by_service(df)
    summary = summarize(df_agg)

    # 把每个 (host, key) 的代表性 comm/cmdline/cwd 取出来做角色推断
    agg_cols = {"comm": "first", "cmdline": "first", "cwd": "first"}
    if "container" in df_agg.columns:
        agg_cols["container"] = "first"
    rep = (df_agg.groupby(["host", "cmdline_key"]).agg(agg_cols).reset_index())
    rep_map = {(r.host, r.cmdline_key): r for r in rep.itertuples()}

    roles, bases = [], []
    for row in summary.itertuples():
        r = rep_map.get((row.host, row.key))
        role, why = infer_role(r.comm, r.cmdline, r.cwd or "",
                               getattr(r, "container", None),
                               host_roles_map.get(row.host, set())) if r else ("未匹配", "")
        roles.append(role)
        bases.append(why)
    summary["role"] = roles
    summary["why"] = bases

    def show(metric: str, label: str, fmt):
        top = summary.sort_values(metric, ascending=False).head(args.top_n)
        print(f"\n{'='*92}\nTOP {args.top_n} —— 按{label}（全主机汇总）\n{'='*92}")
        print(f"{'#':>2} {'角色':<10} {'主机':<15} {'服务(cmdline_key)':<32} {label:>14}")
        print("-" * 92)
        for i, row in enumerate(top.itertuples(), 1):
            print(f"{i:>2} {row.role:<10} {row.host:<15} {row.key[:31]:<32} {fmt(getattr(row, metric)):>14}")

    show("rss_kb_peak", "内存峰值", lambda v: f"{v/1024/1024:.2f} GB")
    show("cpu_pct_peak", "CPU峰值", lambda v: f"{v:.0f} %")
    show("io_total_bytes", "IO总量", lambda v: f"{v/1024/1024/1024:.2f} GB")

    if args.excel:
        export_excel(summary, args.excel, args.top_n)
        print(f"\nExcel 已导出 -> {args.excel}")

    unmatched = summary[summary["role"] == "未匹配"]
    if not unmatched.empty:
        print(f"\n未匹配角色的服务（{len(unmatched)} 个），需人工核对：")
        for row in unmatched.sort_values("rss_kb_peak", ascending=False).head(20).itertuples():
            print(f"  {row.host:<15} {row.key:<30} 内存峰值={row.rss_kb_peak/1024/1024:.2f}GB")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
